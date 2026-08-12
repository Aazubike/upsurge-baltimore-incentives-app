"""
generate_test_data.py

Generates randomized, realistic company submissions and POSTs them to the
LIVE /match/results endpoint of your deployed IncentiveIQ app, so Gemini
actually ranks each one against real programs -- exactly what the sheet
needs to show varied real results for the presentation.

WHAT IT DOES
------------
1. Reads Known_Companies_v2.xlsx (the same raw Salesforce export
   data_loader.py loads, header row 13) and builds "pools" of real values:
   counties, industries, employee counts, addresses+zips, business names.
2. For each synthetic record, it INDEPENDENTLY randomizes each field by
   drawing from those pools (not tied together as one real company) --
   this is what gives you combos like "410 Labs' industry x a totally
   different company's county x a randomly generated employee count,"
   which stress-tests the matcher across way more combinations than the
   564 real companies alone would give you.
3. Revenue has no usable real numbers in the source file (values are a mess
   like ".16 - Pb" or "Pre-revenue"), so revenue is generated from a
   realistic random range instead of sampled.
4. Ownership (mwbe_groups) is deliberately left EMPTY for every record --
   there is no real ownership/MWBE data anywhere in your dataset, so we
   don't fabricate one. This is a known, disclosed data gap.
5. POSTs each record to /match/results with a small thread pool (so it
   doesn't hammer the Gemini API past your rate limit), retries on 429
   with backoff, and writes a CHECKPOINT after every single record --
   so if it crashes or you Ctrl+C partway through, re-running the script
   picks up where it left off instead of starting over.
6. Writes a local manifest CSV of every record it submitted (company name,
   every generated field, submission_id, HTTP status, timestamp). This is
   YOUR OWN local record -- separate from the Google Sheet the live app
   already writes to in the background on every submission. Use this
   manifest to identify/filter out this test batch from the Sheet later
   if you ever need to, since these are synthetic records mixed into
   whatever real Sheet the app is already logging to.

HOW TO RUN
----------
1. Edit the CONFIG block below -- set BASE_URL to your deployed app's URL
   (the onrender.com URL, or your custom domain once that's live).
2. Open Terminal.app, cd to this folder, then:

     python3 generate_test_data.py 100      <- test batch of 100 first
     python3 generate_test_data.py 2000     <- the full run, once the
                                                test batch looks good

   (If you don't pass a number, it defaults to TEST_BATCH_SIZE below.)
3. Watch the terminal output. It prints one line per record as it
   completes, plus a running success/fail count.

WHY A SMALL TEST BATCH FIRST
-----------------------------
This hits your LIVE endpoint -- real Gemini calls, real cost, real writes
to your Google Sheet. Running 100 first confirms the payload shape is
accepted (no 422s) and the matcher behaves as expected, before committing
to 2000.
"""

import csv
import json
import random
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import Lock

import openpyxl
import requests

# ============================== CONFIG ===================================

BASE_URL = "https://incentiveiq.org"
ENDPOINT = "/match/results"

KNOWN_COMPANIES_XLSX = Path(__file__).parent / "Known_Companies_v2.xlsx"

TEST_BATCH_SIZE = 100          # used if you don't pass a number on the command line
TOTAL_RECORDS_DEFAULT = 2000   # the full run size

MAX_WORKERS = 2                # concurrent requests in flight at once.
                                # Start conservative -- check your Gemini
                                # rate limit tier in AI Studio before raising.
REQUEST_TIMEOUT_SECONDS = 120  # Gemini ranking against 1,020 programs is slow --
                                # give it real room, especially on a cold Render start
MAX_RETRIES = 4
BACKOFF_BASE_SECONDS = 3       # retry waits: 3s, 6s, 12s, 24s...

CHECKPOINT_FILE = Path(__file__).parent / "checkpoint.json"
MANIFEST_FILE = Path(__file__).parent / "manifest.csv"

COUNTIES = ["Baltimore City", "Baltimore County", "Anne Arundel", "Harford", "Howard", "Carroll", "Cecil"]
STAGES = ["pre-seed", "seed", "early", "growth", "established"]

RANDOM_SEED = None  # set an int here (e.g. 42) if you want reproducible runs

# --- Cool-down settings ---
# The 100-record test showed a real pattern: the app ran clean for a stretch,
# then broke down under sustained load (empty match data on ~25 of the last
# 26 records). Rather than plow through a repeat of that at 20x scale, this
# tracks consecutive empty-match responses and pauses the whole run when it
# sees a stretch of them, giving the server room to recover.
NO_MATCH_STREAK_THRESHOLD = 5      # consecutive empty-match responses that triggers a cool-down
COOLDOWN_SECONDS = 90               # how long to pause when triggered
DISPATCH_STAGGER_SECONDS = 1.0      # small delay between starting each request, on top of MAX_WORKERS,
                                     # to keep the request rate smoother instead of bursty
CHUNK_SIZE = 250                    # pause briefly between chunks regardless of streaks, just to breathe
CHUNK_PAUSE_SECONDS = 20

# ===========================================================================


def load_pools():
    """Reads the raw Known_Companies_v2.xlsx (header row 13, same as
    data_loader.py's header=12 zero-indexed) and builds independent pools
    of real values to randomize from."""
    wb = openpyxl.load_workbook(KNOWN_COMPANIES_XLSX, data_only=True)
    ws = wb.active
    header = [ws.cell(row=13, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(header) if h}

    names, counties, industries, employee_counts, address_zip_pairs = [], [], [], [], []

    for r in range(14, ws.max_row + 1):
        def cell(col_name):
            col = idx.get(col_name)
            return ws.cell(row=r, column=col).value if col else None

        name = cell("Account Name")
        if not name:
            continue
        names.append(str(name).strip())

        county_raw = cell("County SoT")
        mapped_county = map_county(county_raw)
        if mapped_county:
            counties.append(mapped_county)

        industry_raw = cell("Industry SoT")
        cleaned_industry = clean_industry(industry_raw)
        if cleaned_industry:
            industries.append(cleaned_industry)

        emp_count = parse_employee_count(cell("Number of Employees SoT"))
        if emp_count is not None:
            employee_counts.append(emp_count)

        address_raw = cell("Address SoT")
        zip_code = extract_zip_from_address(address_raw)
        if address_raw and zip_code:
            address_zip_pairs.append((str(address_raw).strip(), zip_code))

    return {
        "names": names,
        "counties": counties or COUNTIES,
        "industries": list(set(industries)) or ["Enterprise Technology"],
        "employee_counts": employee_counts or [5, 15, 50],
        "address_zip_pairs": address_zip_pairs,
    }


def map_county(raw_county) -> str | None:
    """Mirrors main.py's _company_prefill_data substring match against the
    canonical 7-county list."""
    if not raw_county:
        return None
    for c in COUNTIES:
        if c.lower() in str(raw_county).lower():
            return c
    return None


def clean_industry(raw_industry) -> str | None:
    """Mirrors data_loader.get_industry_options()'s per-value cleaning."""
    if not raw_industry:
        return None
    v = str(raw_industry).split(" - ")[0].strip()
    if not v or v.lower() == "no value":
        return None
    return v


def parse_employee_count(value) -> int | None:
    """Exact copy of data_loader.parse_employee_count -- same range-midpoint
    logic, so our randomized employee counts look like real ones."""
    if value is None:
        return None
    s = str(value)
    m = re.search(r"(\d+)\s*-\s*(\d+)", s)
    if m:
        return (int(m.group(1)) + int(m.group(2))) // 2
    m = re.search(r"(\d+)\+", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)", s)
    if m:
        return int(m.group(1))
    return None


def extract_zip_from_address(address) -> str | None:
    """Exact copy of data_loader._extract_zip_from_address."""
    if not address or not isinstance(address, str):
        return None
    matches = re.findall(r"\b(\d{5})\b", address)
    for m in matches:
        if 20600 <= int(m) <= 21999:
            return m
    return matches[0] if matches else None


def random_revenue() -> int:
    """No usable real revenue numbers exist in the source data (see module
    docstring), so this draws from a realistic log-ish spread instead of a
    flat range -- most small/early companies cluster low, with a long tail."""
    bucket = random.choices(
        population=["pre_revenue", "small", "mid", "larger"],
        weights=[0.25, 0.40, 0.25, 0.10],
        k=1,
    )[0]
    if bucket == "pre_revenue":
        return 0
    if bucket == "small":
        return random.randint(10_000, 250_000)
    if bucket == "mid":
        return random.randint(250_000, 2_000_000)
    return random.randint(2_000_000, 20_000_000)


def build_random_submission(pools: dict, index: int) -> dict:
    """Independently randomizes every field from the real-value pools --
    NOT tied together as one real company. Ownership is deliberately left
    empty (see module docstring: no real ownership data exists to sample)."""
    address, zip_code = random.choice(pools["address_zip_pairs"])
    return {
        "company_name": f"{random.choice(pools['names'])} (synthetic-{index})",
        "county": random.choice(pools["counties"]),
        "stage": random.choice(STAGES),
        "industry": random.choice(pools["industries"]),
        "employee_count": str(random.choice(pools["employee_counts"])),
        "annual_revenue": str(random_revenue()),
        "street_address": address,
        "zip_code": zip_code,
        # mwbe_groups intentionally omitted -- no real data to draw from
    }


def load_checkpoint() -> set:
    if CHECKPOINT_FILE.exists():
        return set(json.loads(CHECKPOINT_FILE.read_text()))
    return set()


def response_has_no_matches(html_text: str) -> bool:
    """Checks the actual rendered results page for the 'No matches found' heading
    from results.html. This is real-time signal that the ranking step came back
    empty (the Phase 3 failure mode from the 100-record test), not just a guess
    based on HTTP status -- a 200 response tells us nothing about whether Gemini
    actually returned scored matches."""
    return "No matches found" in html_text


def save_checkpoint(done_indices: set, lock: Lock):
    with lock:
        CHECKPOINT_FILE.write_text(json.dumps(sorted(done_indices)))


def append_manifest_row(row: dict, lock: Lock, write_header: bool):
    with lock:
        file_exists = MANIFEST_FILE.exists()
        with open(MANIFEST_FILE, "a", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            if not file_exists:
                writer.writeheader()
            writer.writerow(row)


def submit_one(session: requests.Session, index: int, payload: dict) -> dict:
    """POSTs one record with retry+backoff on 429/5xx. Returns a result dict
    for the manifest, including whether the results page actually rendered
    matches or came back empty."""
    url = BASE_URL.rstrip("/") + ENDPOINT
    last_error = ""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.post(url, data=payload, timeout=REQUEST_TIMEOUT_SECONDS)
            if resp.status_code == 200:
                no_matches = response_has_no_matches(resp.text)
                return {
                    "index": index, "status": "success", "http_status": 200,
                    "no_matches": no_matches, "error": "", **payload,
                }
            if resp.status_code == 429:
                wait = BACKOFF_BASE_SECONDS * (2 ** (attempt - 1))
                print(f"  [{index}] 429 rate-limited -- waiting {wait}s (attempt {attempt}/{MAX_RETRIES})")
                time.sleep(wait)
                last_error = "429 rate limited"
                continue
            if 500 <= resp.status_code < 600:
                wait = BACKOFF_BASE_SECONDS * (2 ** (attempt - 1))
                print(f"  [{index}] {resp.status_code} server error -- waiting {wait}s (attempt {attempt}/{MAX_RETRIES})")
                time.sleep(wait)
                last_error = f"{resp.status_code} server error"
                continue
            # 4xx that isn't 429 (e.g. 422 validation error) -- retrying won't help
            return {
                "index": index, "status": "failed", "http_status": resp.status_code,
                "no_matches": None, "error": resp.text[:300], **payload,
            }
        except requests.RequestException as e:
            wait = BACKOFF_BASE_SECONDS * (2 ** (attempt - 1))
            print(f"  [{index}] network error ({e}) -- waiting {wait}s (attempt {attempt}/{MAX_RETRIES})")
            time.sleep(wait)
            last_error = str(e)
    return {"index": index, "status": "failed", "http_status": None, "no_matches": None, "error": last_error, **payload}


def main():
    if RANDOM_SEED is not None:
        random.seed(RANDOM_SEED)

    total = int(sys.argv[1]) if len(sys.argv) > 1 else TEST_BATCH_SIZE
    print(f"Generating and submitting {total} synthetic records to {BASE_URL}{ENDPOINT}")
    print(f"Concurrency: {MAX_WORKERS} workers | Checkpoint: {CHECKPOINT_FILE.name}\n")

    if "YOUR-APP-NAME" in BASE_URL:
        print("!! Edit BASE_URL at the top of this script before running. Stopping.")
        return

    pools = load_pools()
    print(f"Loaded pools from {KNOWN_COMPANIES_XLSX.name}: "
          f"{len(pools['names'])} names, {len(set(pools['counties']))} counties, "
          f"{len(pools['industries'])} industries, {len(pools['address_zip_pairs'])} address/zip pairs\n")

    done_indices = load_checkpoint()
    if done_indices:
        print(f"Resuming: {len(done_indices)} records already completed in a previous run.\n")

    remaining = [i for i in range(total) if i not in done_indices]
    if not remaining:
        print("Nothing to do -- all records already completed. Delete checkpoint.json to start fresh.")
        return

    checkpoint_lock = Lock()
    manifest_lock = Lock()
    success_count, fail_count, no_match_count = 0, 0, 0
    consecutive_no_match = 0
    session = requests.Session()

    chunks = [remaining[i:i + CHUNK_SIZE] for i in range(0, len(remaining), CHUNK_SIZE)]

    for chunk_num, chunk in enumerate(chunks, start=1):
        print(f"\n--- Chunk {chunk_num}/{len(chunks)} ({len(chunk)} records) ---")
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {}
            for i in chunk:
                payload = build_random_submission(pools, i)
                futures[executor.submit(submit_one, session, i, payload)] = i
                time.sleep(DISPATCH_STAGGER_SECONDS)  # stagger dispatch so requests don't all fire at once

            for future in as_completed(futures):
                result = future.result()
                index = result["index"]
                done_indices.add(index)
                save_checkpoint(done_indices, checkpoint_lock)

                manifest_row = {
                    "index": index,
                    "status": result["status"],
                    "http_status": result["http_status"],
                    "no_matches": result.get("no_matches"),
                    "error": result["error"],
                    "company_name": result["company_name"],
                    "county": result["county"],
                    "stage": result["stage"],
                    "industry": result["industry"],
                    "employee_count": result["employee_count"],
                    "annual_revenue": result["annual_revenue"],
                    "street_address": result["street_address"],
                    "zip_code": result["zip_code"],
                    "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
                }
                append_manifest_row(manifest_row, manifest_lock, write_header=(not MANIFEST_FILE.exists()))

                if result["status"] == "success":
                    success_count += 1
                    if result.get("no_matches"):
                        no_match_count += 1
                        consecutive_no_match += 1
                        print(f"[{index+1}/{total}] OK, but NO MATCHES rendered -- {result['company_name']} "
                              f"(streak: {consecutive_no_match})")
                    else:
                        consecutive_no_match = 0
                        print(f"[{index+1}/{total}] OK -- {result['company_name']}")
                else:
                    fail_count += 1
                    consecutive_no_match = 0  # a hard failure isn't the same signal as an empty-but-successful page
                    print(f"[{index+1}/{total}] FAIL -- {result['company_name']} -- {result['error']}")

                if consecutive_no_match >= NO_MATCH_STREAK_THRESHOLD:
                    print(f"\n!! {consecutive_no_match} consecutive empty-match responses detected. "
                          f"Cooling down for {COOLDOWN_SECONDS}s before continuing, so the server gets a break "
                          f"instead of running straight through a breakdown like the 100-record test showed.\n")
                    time.sleep(COOLDOWN_SECONDS)
                    consecutive_no_match = 0

        if chunk_num < len(chunks):
            print(f"--- Chunk {chunk_num} done. Pausing {CHUNK_PAUSE_SECONDS}s before the next chunk. ---")
            time.sleep(CHUNK_PAUSE_SECONDS)

    print(f"\nDone. {success_count} succeeded ({no_match_count} of those came back with no matches), {fail_count} failed.")
    print(f"Manifest written to {MANIFEST_FILE}")
    if fail_count or no_match_count:
        print("Re-run the same command to retry only the incomplete records (checkpoint skips completed ones).")
        print("Note: records that succeeded with no matches ARE marked complete in the checkpoint -- "
              "they got a real response, just an empty one. Check manifest.csv's no_matches column to find them "
              "if you want to specifically re-submit those.")


if __name__ == "__main__":
    main()
